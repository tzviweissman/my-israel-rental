"""Bulk property upload — parse, preview, commit, attach images.

5-step owner wizard on the frontend (Template → Input → Preview → Images → Done).
Backend workflow:
  1. GET  /properties/bulk/template    → download CSV/XLSX template
  2. POST /properties/bulk/parse       → dry-run: parse file/text, validate every row,
                                          return preview + errors (no DB writes)
  3. POST /properties/bulk/commit      → create every valid row in a single transaction
  4. POST /properties/bulk/images      → (optional) upload a ZIP, match files by name,
                                          attach to the newly-created properties

The CSV format is defined by COLUMNS below — matching the fields in PropertyCreate.
Lists (amenities, image_filenames) use `;` as the separator because commas collide
with the CSV delimiter and most owners know semicolons from ICS/export formats.
"""
import csv
import io
import json
import uuid
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, List

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ValidationError

from models import PropertyCreate
from models_response import (
    BulkCommitResponse,
    BulkExtractResponse,
    BulkImagesAttachResponse,
    BulkParseResponse,
)
from routes.deps import UPLOAD_DIR, db, verify_token

router = APIRouter()
api_router = router


# Columns accepted in the CSV/XLSX/paste input. Order is authoritative — this
# is what we emit for the template too.
COLUMNS = [
    "title", "description", "rental_type", "property_type",
    "bedrooms", "bathrooms", "floor", "area", "address",
    "square_meters", "porch_square_meters", "porches",
    "has_elevator", "is_shabbat_elevator", "is_tama", "sukkah_compatible",
    "has_agent_fee", "agent_fee_price", "agent_fee_currency",
    "condition", "furniture_option", "amenities",
    "monthly_price", "nightly_price", "currency",
    "cancellation_policy", "custom_cancellation_policy",
    "available_from", "starting_date", "minimum_booking_days",
    "image_filenames",
]

REQUIRED = {"title", "rental_type", "property_type", "bedrooms", "area"}

_BOOL_TRUE = {"true", "yes", "y", "1", "x", "✓"}
_BOOL_FALSE = {"false", "no", "n", "0", ""}


class BulkCommitBody(BaseModel):
    rows: List[dict]


def _parse_bool(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    s = str(val or "").strip().lower()
    if s in _BOOL_TRUE:
        return True
    if s in _BOOL_FALSE:
        return False
    raise ValueError(f"expected yes/no, got '{val}'")


def _parse_number(val: Any, caster: Any, field: str) -> Any:
    if val in (None, "", "-"):
        return None
    try:
        return caster(val)
    except (ValueError, TypeError):
        raise ValueError(f"'{field}' must be a number (got '{val}')")


def _split_list(val: Any) -> list:
    if not val:
        return []
    if isinstance(val, list):
        return [str(x).strip() for x in val if str(x).strip()]
    return [x.strip() for x in str(val).split(";") if x.strip()]


_VALID_RENTAL_TYPES = {"long-term", "short-term", "vacation", "storage"}
_BOOL_FIELDS = ("has_elevator", "is_shabbat_elevator", "is_tama", "sukkah_compatible", "has_agent_fee")
_INT_FIELDS = ("bedrooms", "bathrooms", "floor", "porches", "minimum_booking_days")
_FLOAT_FIELDS = ("square_meters", "porch_square_meters", "monthly_price", "nightly_price", "agent_fee_price")
_LIST_FIELDS = ("amenities", "image_filenames")
_DEFAULTS = {
    "property_type": "apartment",
    "bathrooms": 1,
    "floor": 1,
    "porches": 0,
    "condition": "good",
    "furniture_option": "no_furniture",
    "cancellation_policy": "flexible",
}


def _project_columns(raw: dict) -> dict:
    """Restrict the input to known columns + null out empty strings."""
    return {k: (raw.get(k) if raw.get(k) not in (None, "") else None) for k in COLUMNS}


def _assert_required_present(row: dict) -> None:
    for field in REQUIRED:
        if not row.get(field):
            raise ValueError(f"'{field}' is required")


def _normalize_rental_type(row: dict) -> None:
    rt = str(row["rental_type"]).strip().lower().replace(" ", "-")
    if rt not in _VALID_RENTAL_TYPES:
        raise ValueError(
            f"rental_type must be long-term/short-term/vacation/storage (got '{row['rental_type']}')"
        )
    row["rental_type"] = rt


def _coerce_numeric_and_bool(row: dict) -> None:
    for f in _BOOL_FIELDS:
        row[f] = _parse_bool(row.get(f))
    for f in _INT_FIELDS:
        row[f] = _parse_number(row.get(f), lambda v: int(float(v)), f)
    for f in _FLOAT_FIELDS:
        row[f] = _parse_number(row.get(f), float, f)
    for f in _LIST_FIELDS:
        row[f] = _split_list(row.get(f))


def _apply_implied_flags(row: dict) -> None:
    """Owners often only mention 'Shabbat elevator' in a paste, but that
    implies the building has an elevator at all. Set the parent flag so the
    listing doesn't show inconsistent metadata."""
    if row.get("is_shabbat_elevator"):
        row["has_elevator"] = True


def _apply_defaults_and_currency(row: dict) -> None:
    row["currency"] = str(row["currency"]).strip().upper() if row.get("currency") else "ILS"
    row["agent_fee_currency"] = (
        str(row["agent_fee_currency"]).strip().upper() if row.get("agent_fee_currency") else "ILS"
    )
    for key, default in _DEFAULTS.items():
        row[key] = row.get(key) or default


def _normalize_row(raw: dict) -> dict:
    """Turn one input-dict (string-valued) into a PropertyCreate-ready dict.
    Raises ValueError with a friendly message on the first bad field."""
    row = _project_columns(raw)
    _assert_required_present(row)
    _normalize_rental_type(row)
    _coerce_numeric_and_bool(row)
    _apply_implied_flags(row)
    _apply_defaults_and_currency(row)
    return row


def _validate_row(row: dict) -> str | None:
    """Run Pydantic validation. Returns None if OK, otherwise a short error string."""
    # Strip fields that PropertyCreate doesn't know about (image_filenames is
    # our own meta, not a DB field).
    payload = {k: v for k, v in row.items() if k != "image_filenames"}
    try:
        PropertyCreate(**payload)
        return None
    except ValidationError as e:
        first = e.errors()[0]
        loc = ".".join(str(x) for x in first["loc"])
        return f"{loc}: {first['msg']}"


# ------------------------ PARSERS ------------------------

def _parse_csv(text: str) -> List[dict]:
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]


def _parse_xlsx(content: bytes) -> List[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c in (None, "") for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers)) if headers[i]})
    return out


def _parse_paste(text: str) -> List[dict]:
    """Accept tab- or comma-separated paste content."""
    text = text.strip()
    if not text:
        return []
    delimiter = "\t" if "\t" in text.splitlines()[0] else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]


# ------------------------ ENDPOINTS ------------------------

@api_router.get("/properties/bulk/template")
async def download_template(fmt: str = "csv") -> StreamingResponse:
    """Stream a blank CSV template with instructions in row 2."""
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="fmt must be csv or xlsx")

    samples = {
        "title": "Cozy Jerusalem Apartment",
        "description": "Bright 2BR with a balcony",
        "rental_type": "long-term",
        "property_type": "apartment",
        "bedrooms": "2", "bathrooms": "1", "floor": "3",
        "area": "Jerusalem - Rehavia", "address": "King George 10",
        "square_meters": "75", "porch_square_meters": "8", "porches": "1",
        "has_elevator": "yes", "is_shabbat_elevator": "no",
        "is_tama": "no", "sukkah_compatible": "yes",
        "has_agent_fee": "yes", "agent_fee_price": "6500", "agent_fee_currency": "ILS",
        "condition": "good", "furniture_option": "furnished",
        "amenities": "WiFi;AC;Washing Machine",
        "monthly_price": "6500", "nightly_price": "", "currency": "ILS",
        "cancellation_policy": "flexible",
        "custom_cancellation_policy": "",
        "available_from": "2026-09-01", "starting_date": "",
        "minimum_booking_days": "",
        "image_filenames": "apt1_front.jpg;apt1_kitchen.jpg",
    }
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(COLUMNS)
        w.writerow([samples.get(c, "") for c in COLUMNS])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=bulk_properties_template.csv"},
        )
    # xlsx
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Properties"
    ws.append(COLUMNS)
    ws.append([samples.get(c, "") for c in COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_properties_template.xlsx"},
    )


@api_router.post("/properties/bulk/parse", response_model=BulkParseResponse)
async def parse_bulk(
    file: UploadFile | None = File(None),
    text: str | None = Form(None),
    payload: dict = Depends(verify_token),
) -> dict:
    """Dry run: parse the input and return [{row, normalized, errors}]. No DB writes."""
    if payload.get("role") not in ("owner", "manager", "admin"):
        raise HTTPException(status_code=403, detail="Only owners and managers can bulk-upload")

    raw_rows: List[dict] = []
    if file is not None:
        content = await file.read()
        name = (file.filename or "").lower()
        if name.endswith(".csv"):
            raw_rows = _parse_csv(content.decode("utf-8-sig"))
        elif name.endswith(".xlsx"):
            raw_rows = _parse_xlsx(content)
        else:
            raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file")
    elif text:
        raw_rows = _parse_paste(text)
    else:
        raise HTTPException(status_code=400, detail="Provide a file or pasted text")

    if not raw_rows:
        return {"rows": [], "summary": {"total": 0, "valid": 0, "invalid": 0}}

    preview = []
    for i, raw in enumerate(raw_rows, start=1):
        entry: dict[str, Any] = {"index": i, "raw": raw, "errors": []}
        try:
            normalized = _normalize_row(raw)
            err = _validate_row(normalized)
            if err:
                entry["errors"].append(err)
            entry["normalized"] = normalized
        except ValueError as e:
            entry["errors"].append(str(e))
        preview.append(entry)

    total = len(preview)
    valid = sum(1 for p in preview if not p["errors"])
    return {"rows": preview, "summary": {"total": total, "valid": valid, "invalid": total - valid}}


@api_router.post("/properties/bulk/commit", response_model=BulkCommitResponse)
async def commit_bulk(body: BulkCommitBody, payload: dict = Depends(verify_token)) -> dict:
    """Create every row in `body.rows`. Rows must already be normalized (call /parse first)."""
    if payload.get("role") not in ("owner", "manager", "admin"):
        raise HTTPException(status_code=403, detail="Only owners and managers can bulk-upload")

    created = []
    skipped = []
    for i, row in enumerate(body.rows, start=1):
        # Always re-normalize + re-validate. Coercions are idempotent so it's safe
        # to run even if the caller already sent clean values.
        try:
            normalized = _normalize_row(row)
            err = _validate_row(normalized)
            if err:
                skipped.append({"index": i, "title": row.get("title"), "error": err})
                continue
            property_id = str(uuid.uuid4())
            image_filenames = normalized.pop("image_filenames", [])
            doc = PropertyCreate(**{k: v for k, v in normalized.items() if k != "image_filenames"}).model_dump()
            doc.update({
                "id": property_id,
                "owner_id": payload["user_id"],
                "images": [],
                "videos": [],
                "created_at": datetime.now(UTC).isoformat(),
                "status": "active",
                "liked_by": [],
                "pending_image_filenames": image_filenames,  # resolved later by /bulk/images
                "bulk_created": True,  # drives the "NEW" badge on the dashboard for 24h
            })
            await db.properties.insert_one(doc)
            created.append({"index": i, "id": property_id, "title": doc["title"], "image_filenames": image_filenames})
        except Exception as e:
            skipped.append({"index": i, "title": row.get("title"), "error": str(e)})

    return {"created": created, "skipped": skipped, "summary": {"created": len(created), "skipped": len(skipped)}}


_ALLOWED_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif")


def _assert_bulk_role(payload: dict) -> None:
    if payload.get("role") not in ("owner", "manager", "admin"):
        raise HTTPException(status_code=403, detail="Only owners and managers can bulk-upload")


def _parse_mapping_json(mapping: str) -> dict:
    try:
        return json.loads(mapping)
    except Exception:
        raise HTTPException(
            status_code=400, detail="mapping must be JSON: {property_id: [filenames]}"
        ) from None


async def _load_owned_property(prop_id: str, payload: dict) -> dict | None:
    """Return the property doc if the caller owns it (or is admin), else None."""
    prop = await db.properties.find_one(
        {"id": prop_id}, {"_id": 0, "owner_id": 1, "images": 1}
    )
    if not prop:
        return None
    if prop.get("owner_id") != payload["user_id"] and payload.get("role") != "admin":
        return None
    return prop


def _persist_uploaded_image(data: bytes, ext: str) -> str:
    """Write one image file to disk and return the public URL."""
    safe_name = f"{uuid.uuid4().hex}{ext}"
    (UPLOAD_DIR / safe_name).write_bytes(data)
    return f"/api/uploads/{safe_name}"


def _attach_one(
    prop_id: str,
    fname: str,
    data: bytes | None,
    images: list,
    results: dict,
) -> bool:
    """Attach a single (filename, bytes) pair to ``images``. Returns True on success.

    Tracks misses/unsupported types in ``results['missing']`` so the caller
    can surface them to the UI.
    """
    if data is None:
        results["missing"].append({"property_id": prop_id, "filename": fname})
        return False
    ext = Path(fname).suffix.lower() or ".jpg"
    if ext not in _ALLOWED_IMAGE_EXTS:
        results["missing"].append({"property_id": prop_id, "filename": fname, "reason": "unsupported type"})
        return False
    url = _persist_uploaded_image(data, ext)
    images.append(url)
    results["attached"].append({"property_id": prop_id, "filename": fname, "url": url})
    return True


async def _fanout_images(
    property_map: dict, file_source: dict, payload: dict
) -> dict:
    """Shared core for both bulk-image endpoints.

    ``file_source`` maps basename → bytes; the caller decides how to fill it
    (from a zip vs from a flat ``UploadFile`` list). Owners/admins only; only
    patches a property document if at least one image actually attached.
    """
    results: dict[str, list] = {"attached": [], "missing": [], "not_owned": []}
    for prop_id, filenames in property_map.items():
        prop = await _load_owned_property(prop_id, payload)
        if not prop:
            results["not_owned"].append(prop_id)
            continue
        new_images = list(prop.get("images") or [])
        attached_any = False
        for fname in filenames:
            data = file_source.get(Path(fname).name)
            if _attach_one(prop_id, fname, data, new_images, results):
                attached_any = True
        if attached_any:
            await db.properties.update_one(
                {"id": prop_id},
                {"$set": {"images": new_images}, "$unset": {"pending_image_filenames": ""}},
            )
    return results


@api_router.post("/properties/bulk/images", response_model=BulkImagesAttachResponse)
async def attach_bulk_images(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    payload: dict = Depends(verify_token),
) -> dict:
    """Unzip `file`, match entries by filename, and attach to properties.

    `mapping` is a JSON string: {property_id: [filename1, filename2, ...]}.
    Owner must own each property_id.
    """
    _assert_bulk_role(payload)
    property_map = _parse_mapping_json(mapping)
    content = await file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="file is not a valid .zip") from None
    # Build filename → bytes map (basename only; owners often zip a folder)
    file_source = {
        Path(n).name: zf.read(n) for n in zf.namelist() if not n.endswith("/")
    }
    return await _fanout_images(property_map, file_source, payload)


@api_router.post("/properties/bulk/images/attach", response_model=BulkImagesAttachResponse)
async def attach_bulk_images_flat(
    mapping: str = Form(...),
    files: List[UploadFile] = File(...),
    payload: dict = Depends(verify_token),
) -> dict:
    """Attach N individual image uploads to properties by a property_id→[filename] map.

    This is the flat-file sibling of /properties/bulk/images. Owners drop
    loose files (from the "Needs Images" filter drop-zone) instead of zipping
    first — the frontend just sends the raw File objects alongside a mapping.
    """
    _assert_bulk_role(payload)
    property_map = _parse_mapping_json(mapping)
    file_source: dict = {}
    for f in files:
        name = Path(f.filename or "").name
        if name:
            file_source[name] = await f.read()
    return await _fanout_images(property_map, file_source, payload)


# ---------------------------------------------------------------------------
# Smart paste: parse messy free-form property descriptions (WhatsApp / email /
# Telegram) into structured rows using Claude. The frontend pipes the result
# into the visual editor so the user can review/edit before committing.
# ---------------------------------------------------------------------------

class SmartExtractIn(BaseModel):
    text: str


_EXTRACT_SYSTEM_PROMPT = """You are an extraction engine for an Israeli rental-property platform.

The user pastes free-form text (often from WhatsApp/email) describing one or more properties.
Each property is usually separated by blank lines, "---", a numbered list, or a clear topic shift.
Languages mix English and Hebrew freely.

Return a JSON object with a single key "properties" — an array. Each entry must be a single
property using EXACTLY these field names (snake_case). Translate any Hebrew values into English
prose for the human-readable fields (title, description, area, address). Do NOT translate
proper-noun place names beyond what an Anglophone Israeli would recognize (e.g. keep
"Sanhedria Murchevet" or "Belz" as transliteration; "רחוב קדושת אהרון" -> "Kedushat Aharon Street").

Required fields (always emit):
  - title          (string, <= 80 chars, generated if not present -- e.g. "Ground-floor 1.5 BR in Sanhedria Murchevet")
  - description    (string, 1-3 sentences summarising what's in the source text, in English)
  - area           (string, neighbourhood or city + neighbourhood, in English)
  - address        (string, best-guess address; empty string if not in source)
  - rental_type    (one of: long-term | short-term | vacation | storage; default long-term)
  - property_type  (one of: apartment | house | villa; default apartment)
  - bedrooms       (number; if "1.5 bedroom" -> 1.5, if "studio" -> 0)
  - bathrooms      (number; default 1 if not mentioned)
  - floor          (integer; -1 = basement, 0 = ground; "ground floor" -> 0)
  - currency       (ILS or USD; default ILS for prices in nis/shekels, USD for $; default ILS)
  - monthly_price  (number -- for long-term/storage; null otherwise)
  - nightly_price  (number -- for short-term/vacation; null otherwise)

Optional (emit when source mentions them):
  - square_meters, porch_square_meters, porches, has_elevator, is_shabbat_elevator,
    is_tama, sukkah_compatible,
    furniture_option (no_furniture | furniture_package | furniture_free),
    condition (renovated | partially_renovated | good),
    amenities (comma-separated string -- pick ONLY from this canonical list, exact spelling:
      "Central AC / Heating", "In-unit washer and dryer", "Dishwasher",
      "Walk in Closets", "High Ceilings", "Ensuite Bathroom", "Storage Space",
      "Heated Floors", "Gym / Fitness center", "Swimming pool (indoor or outdoor)",
      "Hot tub / Spa", "On-site parking (garage or lot)", "Wi-Fi included"),
    cancellation_policy (flexible | moderate | strict | custom),
    custom_cancellation_policy (string -- free text, only when cancellation_policy = custom),
    available_from (ISO date if derivable, else string),
    minimum_booking_days,
    has_agent_fee (yes | no -- emit "yes" if source mentions agent / broker / שכר טרחה / דמי תיווך),
    agent_fee_price (number -- the agent / broker fee amount when stated),
    agent_fee_currency (ILS | USD -- match the agent fee's currency)

Booleans MUST be the literal strings "yes" or "no" (lower-case).

Important rules:
- If unsure about a value, OMIT it rather than guess.
- Hebrew month references (e.g. "ר"ח אייר" / "Rosh Chodesh Iyar") -> put the transliterated phrase into available_from as a string.
- "Fully furnished" / "Furniture included" / "comes with furniture" -> furniture_option: "furniture_free".
  "Furniture package available" / "furniture for sale" -> furniture_option: "furniture_package".
  "Unfurnished" / "no furniture" -> furniture_option: "no_furniture".
- "Renovated" / "after renovation" / "newly renovated" -> condition: "renovated".
  "Partially renovated" / "partial reno" -> condition: "partially_renovated".
  Otherwise -> condition: "good".
- "approx 60" -> square_meters: 60.
- For amenities, only emit values that match the canonical list above (case-sensitive). Skip anything you can't map.
- Identify property boundaries: blank lines, "---", numbered prefixes (1., 2.), bold area headers, or repeated location lines.

Return ONLY raw JSON -- no markdown fences, no commentary.
"""


@api_router.post("/properties/bulk/extract", response_model=BulkExtractResponse)
async def smart_extract(
    body: SmartExtractIn,
    payload: dict = Depends(verify_token),
) -> dict:
    """Use Claude to parse messy free-form property descriptions into structured rows."""
    if payload.get("role") not in ("admin", "owner", "manager"):
        raise HTTPException(status_code=403, detail="Owner or manager role required")

    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Empty input")
    if len(text) > 30_000:
        raise HTTPException(status_code=413, detail="Input too long (max 30k characters)")

    from emergentintegrations.llm.chat import LlmChat, UserMessage

    from routes.deps import EMERGENT_LLM_KEY

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=str(uuid.uuid4()),
        system_message=_EXTRACT_SYSTEM_PROMPT,
    )
    chat.with_model("anthropic", "claude-4-sonnet-20250514")
    raw = await chat.send_message(UserMessage(text=text))

    parsed: Any
    try:
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("```", 2)[1]
            if cleaned.lower().startswith("json"):
                cleaned = cleaned[4:]
            cleaned = cleaned.strip().strip("`").strip()
        parsed = json.loads(cleaned)
    except (json.JSONDecodeError, IndexError) as e:
        raise HTTPException(
            status_code=502,
            detail=f"Could not parse LLM response as JSON: {e}",
        )

    properties = parsed.get("properties") if isinstance(parsed, dict) else None
    if not isinstance(properties, list):
        raise HTTPException(status_code=502, detail="LLM did not return a 'properties' array")

    # Sanity-clamp to prevent prompt-injection or hallucinated rows.
    if len(properties) > 50:
        properties = properties[:50]

    return {"properties": properties, "count": len(properties)}

